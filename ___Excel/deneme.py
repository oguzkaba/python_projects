from openpyxl import Workbook,load_workbook
from openpyxl.drawing.image import Image
import pyqrcode as pc

# generate QR
a=pc.create("GV-PR-01502")
a.png("../fakedata/b.png",scale=3,quiet_zone=0)
qr_image=Image("../fakedata/b.png")
# QR add Exel file
wb = load_workbook("../fakedata/GV-PR-01502.xlsx")
ws = wb.active
# # Aktif çalışma sayfasının adını yazdırma
# print(wb.sheetnames)      # <Worksheet "İsimler">

# # for satir in range(2,5):    
# #     for sutun in range(1,3):        
# #         print(" | " + str(ws.cell(satir,sutun).value) + " | ",end="")
# #     print()

# # for satir in ws['A2':'b4']:
# #     for hucre in satir:
# #         print(" | " + str(hucre.value) + " | ",end="")
# #     print() 

# wb = Workbook()

# ws = wb.active
# ws.title = "İlk Çalışma Alanı"
# ws = wb.create_sheet("Posta Kodları")
# ws = wb.create_sheet("Ülkeler")
# try:
#     wb.save("../fakedata/dosyaAdi.xlsx")
# except:
#     print("hata")
#     wb.close()
# print(wb.sheetnames)   

# sheets=wb.sheetnames
# ws=wb[sheets[0]]
# ws['A49'] = 42
# ws['B3'] = "Merhaba"
# ws.append([1, 2, 3, "Merhaba", "Dünya"])  # Sıradaki satıra sırasıyla dizi elemanlarını ekler

# 1 pixel= 0.026458333 cm

qr_image.width=40
qr_image.height=40
ws.add_image(qr_image,"A49")

try:
    wb.save("../fakedata/GV-PR-01502.xlsx")
except:
    print("HATA-Dosya açık olabilir..!")
    wb.close()
print(wb.sheetnames)   